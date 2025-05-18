import tkinter as tk
import cv2
import openpyxl
import os
import  platform

from tkinter import ttk
from PIL import Image, ImageTk
from datetime import datetime
from EvAQr import abrir_Ventana_qr 

#Funcionalidad
anterior = None

#Crear o cargar el archivo Excel
def crear_o_cargar_excel(nombre_archivo):
    try:
        workbook = openpyxl.load_workbook(nombre_archivo)
        hoja = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        hoja = workbook.active
        hoja.append(["Codigo Qr", "Carrera"])
        workbook.save(nombre_archivo)
    return workbook, hoja

#Verifica si existe la Columna con la fecha actual si no para añadirla
def obtener_columna_fecha(hoja, fecha):
    for col in range(3, hoja.max_column + 1):
        if hoja.cell(row=1, column=col).value == fecha:
            return col
    nueva_columna = hoja.max_column + 1
    hoja.cell(row=1, column=nueva_columna).value = fecha
    return nueva_columna

#Registrar la asistencia con la fecha correspondiente
def registrar_asistencia(nombre_archivo, qr_data):
    workbook, hoja = crear_o_cargar_excel(nombre_archivo)
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    columna_fecha = obtener_columna_fecha(hoja, fecha_actual)
    global anterior

    # Separar el código QR y la carrera
    if " - " in qr_data:
        codigo, carrera = qr_data.split(" - ", 1)
    else:
        codigo = qr_data
        carrera = "No especificada"

    for fila in range(2, hoja.max_row + 1):
        if hoja.cell(row=fila, column=1).value == codigo:
            hoja.cell(row=fila, column=2).value = carrera
            hoja.cell(row=fila, column=columna_fecha).value = "Presente"
            break
    else:
        nueva_fila = hoja.max_row + 1
        hoja.cell(row=nueva_fila, column=1).value = codigo
        hoja.cell(row=nueva_fila, column=2).value = carrera
        hoja.cell(row=nueva_fila, column=columna_fecha).value = "Presente"

    workbook.save(nombre_archivo)

    if anterior is not None and qr_data == anterior:
        print(f"Asistencia registrada: {codigo} ({carrera}) en la fecha {fecha_actual}")

    anterior = qr_data

#Abrir excel
def abrir_excel():
    nombre_archivo = "asistencia.xlsx"
    sistema = platform.system()


    if sistema == "Windows":
        os.startfile(nombre_archivo)
    elif sistema == "Darwin":  # macOS
        os.system(f"open {nombre_archivo}")
    else:  # Linux
        os.system(f"xdg-open {nombre_archivo}")

#Interfaz gráfica

root = tk.Tk()
root.title("EvA - Registro de Asistencia")
root.geometry("1500x850")
#root.attributes("-fullscreen", True)

# Encabezado
header = tk.Frame(root, bg="gray", height=50)
header.pack(fill="x")

tk.Label(header, text="Eva", bg="gray", fg="black", font=("Arial", 16, "bold")).pack(side="left", padx=10)
tk.Button(header, text="❌ Salir", bg="red", fg="white", command=root.quit).pack(side="right")

# Sección de cámara
camara = tk.Frame(root, bg="black", width=400, height=300)
camara.pack(pady=(10, 0), side="left", padx=10)

label_video = tk.Label(camara, bg="black")
label_video.pack()

# Tabla de asistencia
columnas = ("Nombre Completo", "Fecha", "Hora")
TablaAsistencia = ttk.Treeview(root, columns=columnas, show="headings", height=15)
for col in columnas:
    TablaAsistencia.heading(col, text=col)
    TablaAsistencia.column(col, anchor="center")
TablaAsistencia.pack(side="left", padx=(10, 0), fill="both", expand=True)

# Botones
boton_Frame = tk.Frame()
boton_Frame.pack(side="right", padx=(10, 0), fill='y')

tk.Button(boton_Frame, text="Abrir Excel", command = abrir_excel).pack(pady=5)
tk.Button(boton_Frame, text="Crear Qr", command=abrir_Ventana_qr).pack(pady=5)

#Cámara y detección
cap = cv2.VideoCapture(0)
detector = cv2.QRCodeDetector()

def actualizar_video():
    global anterior

    ret, frame = cap.read()
    if not ret:
        print("Sin señal")
        return

    frame = cv2.flip(frame, 1)

    qr_data, bbox, _ = detector.detectAndDecode(frame)

    if qr_data and qr_data != anterior:
        registrar_asistencia("asistencia.xlsx", qr_data)
        fecha_actual = datetime.now().strftime("%Y-%m-%d")
        hora_actual = datetime.now().strftime("%H:%M:%S")
        TablaAsistencia.insert("", "end", values=(qr_data, fecha_actual, hora_actual))
        anterior = qr_data

    # Convertir la imagen a formato compatible con Tkinter
    frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    img = Image.fromarray(frame_rgb)
    imgtk = ImageTk.PhotoImage(image=img)

    label_video.imgtk = imgtk
    label_video.configure(image=imgtk)

    #Actualizacion de 10 milisegundos
    root.after(10, actualizar_video)

# Iniciar el escaneo
actualizar_video()

#Ejecutar
root.mainloop()
cap.release()
cv2.destroyAllWindows()
